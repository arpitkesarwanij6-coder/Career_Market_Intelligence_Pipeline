"""
promote_skills.py
------------------
Run this manually, whenever you want, to turn skills you've clicked [Add] on
(in the web review UI) into real tracked Yes/No/Unknown columns in the master
workbook.

This is deliberately a SEPARATE, manual step from the web app — nothing in
app.py ever changes the workbook's column structure automatically. You stay
in control of when (and whether) the sheet grows.

What it does, every time you run it:
  1. Backs up the master workbook first (same pattern as the app's /api/approve).
  2. Reads skill_dictionary.json's "confirmed_for_next_batch" list.
  3. Appends one new column per confirmed skill, strictly AFTER the current
     last column — it never inserts in the middle, so no existing column
     position, formula, or cross-sheet reference ever shifts.
  4. Copies the header style from the existing skill columns for consistency.
  5. Moves each promoted skill from "confirmed_for_next_batch" into "known"
     in skill_dictionary.json, so the AI extraction schema in app.py picks
     it up as a real fixed skill going forward (restart the server after
     running this).

It does NOT retroactively fill in the new column for job rows you already
approved — those stay blank/Unknown for the new skill until you re-review
them, exactly like any other schema addition.

Usage:
    python promote_skills.py            # promote everything confirmed
    python promote_skills.py --dry-run  # show what would happen, change nothing
"""
import sys
import json
import shutil
from datetime import datetime
from pathlib import Path
from copy import copy

import openpyxl
from openpyxl.utils import get_column_letter

BASE = Path(__file__).resolve().parent
MASTER = BASE / "Career_Market_Intelligence_V4.xlsx"
BACKUPS = BASE / "backups"
SKILL_DICT_PATH = BASE / "skill_dictionary.json"
SHEET_NAME = "01 - Job Database"
HEADER_ROW = 2

def load_dict():
    if not SKILL_DICT_PATH.exists():
        print("No skill_dictionary.json found — nothing to promote.")
        sys.exit(0)
    return json.loads(SKILL_DICT_PATH.read_text(encoding="utf-8"))

def save_dict(d):
    SKILL_DICT_PATH.write_text(json.dumps(d, ensure_ascii=False, indent=2), encoding="utf-8")

def main():
    dry_run = "--dry-run" in sys.argv
    d = load_dict()
    to_promote = d.get("confirmed_for_next_batch", [])

    if not to_promote:
        print("Nothing queued in 'confirmed_for_next_batch'. Nothing to do.")
        return

    print(f"Skills queued for promotion ({len(to_promote)}):")
    for s in to_promote:
        print(f"  - {s}")

    if dry_run:
        print("\n--dry-run: no files were changed.")
        return

    if not MASTER.exists():
        print(f"ERROR: master workbook not found at {MASTER}")
        sys.exit(1)

    BACKUPS.mkdir(exist_ok=True)
    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    backup_path = BACKUPS / f"Career_Market_Intelligence_V4_preworkbook_{stamp}.xlsx"
    shutil.copy2(MASTER, backup_path)
    print(f"\nBackup created: {backup_path.name}")

    wb = openpyxl.load_workbook(MASTER)
    ws = wb[SHEET_NAME]
    start_col = ws.max_column + 1
    template = ws.cell(HEADER_ROW, 24)  # style of an existing skill header (col 24 = "C++")

    for i, skill in enumerate(to_promote):
        col = start_col + i
        cell = ws.cell(HEADER_ROW, col, skill)
        cell.font = copy(template.font)
        cell.fill = copy(template.fill)
        cell.alignment = copy(template.alignment)
        cell.border = copy(template.border)
        ws.column_dimensions[get_column_letter(col)].width = 12
        print(f"  Added column {col} ({get_column_letter(col)}): {skill}")

    wb.save(MASTER)
    print(f"\nSaved. Workbook now has {ws.max_column} columns (was {start_col - 1}).")

    d["known"] = sorted(set(d.get("known", [])) | set(to_promote))
    d["confirmed_for_next_batch"] = []
    save_dict(d)
    print("skill_dictionary.json updated — these skills now count as 'known'.")
    print("\nRestart uvicorn so app.py picks up the updated skill dictionary.")
    print("Note: existing rows will show blank/Unknown for the new column(s)")
    print("until you re-review those postings — this does not back-fill old data.")

if __name__ == "__main__":
    main()
