import os, re, json, shutil, uuid
from datetime import datetime
from pathlib import Path
from urllib.parse import urlparse
import requests
from bs4 import BeautifulSoup
from fastapi import FastAPI, HTTPException
from fastapi.responses import HTMLResponse, FileResponse
from pydantic import BaseModel
from dotenv import load_dotenv
import openpyxl

load_dotenv()
BASE = Path(__file__).resolve().parent
MASTER = BASE / "Career_Market_Intelligence_V4.xlsx"
BACKUPS = BASE / "backups"
STAGING = BASE / "staging"
BACKUPS.mkdir(exist_ok=True)
STAGING.mkdir(exist_ok=True)

app = FastAPI(title="Career Market Intelligence Pipeline V0.1")

class AnalyzeRequest(BaseModel):
    url: str

class ApproveRequest(BaseModel):
    record: dict

class SkillDictAction(BaseModel):
    skill: str

def clean_text(html):
    soup = BeautifulSoup(html, "html.parser")
    for x in soup(["script", "style", "noscript", "svg"]):
        x.decompose()
    text = soup.get_text("\n", strip=True)
    text = re.sub(r"\n{3,}", "\n\n", text)
    return text[:60000]

def fetch(url):
    r = requests.get(url, timeout=25, headers={
        "User-Agent": "Mozilla/5.0 CareerMarketIntelligence/0.1"
    })
    r.raise_for_status()
    return r.url, clean_text(r.text)

def source_type(url):
    host = urlparse(url).netloc.lower()
    if "amazon.jobs" in host: return "Company Careers"
    if "wellfound.com" in host: return "Wellfound"
    if "linkedin.com" in host: return "LinkedIn"
    if "indeed." in host: return "Indeed"
    if "naukri.com" in host: return "Naukri"
    return "Other"

def extract_basic(text, url):
    # Fallback extractor used when no AI key is configured.
    title = ""
    m = re.search(r"(?:^|\n)#?\s*([^\n]{3,120})", text)
    if m: title = m.group(1).strip()
    return {
        "Job ID": "",
        "Date Found": datetime.now().strftime("%Y-%m-%d"),
        "Company": "",
        "Role Title (Raw)": title,
        "Standardized Role": "",
        "Level": "",
        "Country": "",
        "City": "",
        "Location Type": "",
        "Visa Sponsorship": "Unknown",
        "Relocation Support": "Unknown",
        "Currency": "",
        "Salary Min": "",
        "Salary Max": "",
        "Discovery URL": url,
        "Notes": "",
        "Exclude From Analysis": "No",
        "Job Source": source_type(url),
        "Original / Official Source": "Yes" if source_type(url) == "Company Careers" else "Unknown",
        "Official Job URL": url if source_type(url) == "Company Careers" else "",
        "Data Source / Evidence": "Fetched job page; AI not configured.",
        "_confidence": "Low",
        "_evidence": {},
        "_raw_text": text[:8000],
    }

V4_FIELDS = [
"Job ID","Date Found","Company","Role Title (Raw)","Standardized Role","Level","Country","City","Location Type",
"Visa Sponsorship","Relocation Support","Currency","Salary Min","Salary Max","Discovery URL","Notes","Exclude From Analysis",
"C++","Python","Java","Go","Rust","JavaScript/TypeScript","C","SQL","Linux","Operating Systems",
"Multithreading & Concurrency","Networking","Distributed Systems","Compilers","AWS","Azure","GCP","Kubernetes","Docker",
"Terraform","CI/CD","Data Structures & Algorithms","Machine Learning","Deep Learning","PyTorch/TensorFlow","LLMs / NLP",
"CUDA / GPU Programming","Big Data (Spark/Hadoop)","System Design","Microservices / API Design","Security / Cryptography",
"Testing / QA","Base Comp Min","Base Comp Max","Bonus Min (Annual)","Bonus Max (Annual)",
"Equity Min (Annual, Amortized)","Equity Max (Annual, Amortized)",
"Visa Sponsorship Evidence / Source","Relocation Evidence / Source","Job Source",
"Intl Candidate Viability (India->USA)","Visa Sponsorship Data Type","Relocation Data Type",
"Source Type","Original / Official Source","Official Job URL","Data Source / Evidence",
"Additional Skills Detected (Unmapped)"
]

# ---------------------------------------------------------------------------
# Master Skill Dictionary — lives outside the workbook entirely, on purpose.
# "known"   = skills with a real Yes/No/Unknown column in the workbook today.
# "pending" = {skill_name: times_seen} — surfaced by AI, not yet promoted.
# "ignored" = skill names the user has explicitly said not to track.
# Promotion to a real workbook column is a separate, deliberate step
# (see promote_skills.py) — this file never causes the workbook to change.
# ---------------------------------------------------------------------------
SKILL_DICT_PATH = BASE / "skill_dictionary.json"

def load_skill_dict():
    if SKILL_DICT_PATH.exists():
        try:
            d = json.loads(SKILL_DICT_PATH.read_text(encoding="utf-8"))
        except Exception:
            d = {}
    else:
        d = {}
    d.setdefault("known", sorted(BOOL_SKILLS))
    d.setdefault("pending", {})
    d.setdefault("ignored", [])
    d.setdefault("confirmed_for_next_batch", [])
    return d

def save_skill_dict(d):
    SKILL_DICT_PATH.write_text(json.dumps(d, ensure_ascii=False, indent=2), encoding="utf-8")

def register_detected_skills(names):
    """Called after every analyze(). Adds/increments unseen skill names into
    'pending'. Never touches the workbook. Returns the filtered list actually
    surfaced to the reviewer (known + ignored + already-confirmed are dropped)."""
    d = load_skill_dict()
    known_lower = {s.lower() for s in d["known"]}
    ignored_lower = {s.lower() for s in d["ignored"]}
    confirmed_lower = {s.lower() for s in d["confirmed_for_next_batch"]}
    surfaced = []
    for raw in names:
        name = (raw or "").strip()
        if not name or len(name) > 60:
            continue
        low = name.lower()
        if low in known_lower or low in ignored_lower or low in confirmed_lower:
            continue
        d["pending"][name] = d["pending"].get(name, 0) + 1
        surfaced.append(name)
    save_skill_dict(d)
    return surfaced



BOOL_SKILLS = {
    "C++","Python","Java","Go","Rust","JavaScript/TypeScript","C","SQL","Linux","Operating Systems",
    "Multithreading & Concurrency","Networking","Distributed Systems","Compilers","AWS","Azure","GCP",
    "Kubernetes","Docker","Terraform","CI/CD","Data Structures & Algorithms","Machine Learning",
    "Deep Learning","PyTorch/TensorFlow","LLMs / NLP","CUDA / GPU Programming","Big Data (Spark/Hadoop)",
    "System Design","Microservices / API Design","Security / Cryptography","Testing / QA"
}

def ai_extract(text, url):
    key = os.getenv("GEMINI_API_KEY")
    if not key:
        return extract_basic(text, url)

    model = os.getenv("GEMINI_MODEL", "gemini-3.5-flash-lite")
    GEMINI_URL = f"https://generativelanguage.googleapis.com/v1beta/models/{model}:generateContent"

    schema = { 
        "type": "OBJECT",
        "properties": {
            "Job ID": {"type": "STRING"},
            "Date Found": {"type": "STRING"},
            "Company": {"type": "STRING"},
            "Role Title (Raw)": {"type": "STRING"},
            "Standardized Role": {"type": "STRING"},
            "Level": {"type": "STRING"},
            "Country": {"type": "STRING"},
            "City": {"type": "STRING"},
            "Location Type": {"type": "STRING"},
            "Visa Sponsorship": {
                "type": "STRING",
                "enum": ["Yes", "No", "Unknown"]
            },
            "Relocation Support": {
                "type": "STRING",
                "enum": ["Yes", "No", "Unknown"]
            },
            "Currency": {"type": "STRING"},
            "Salary Min": {"type": "STRING"},
            "Salary Max": {"type": "STRING"},
            "Discovery URL": {"type": "STRING"},
            "Notes": {"type": "STRING"},
            "Exclude From Analysis": {"type": "STRING"},

            "skills": {
                "type": "OBJECT",
                "properties": {
                    skill: {
                        "type": "STRING",
                        "enum": ["Yes", "No", "Unknown"]
                    }
                    for skill in BOOL_SKILLS
                },
                "required": list(BOOL_SKILLS)
            },

            "Base Comp Min": {"type": "STRING"},
            "Base Comp Max": {"type": "STRING"},
            "Bonus Min (Annual)": {"type": "STRING"},
            "Bonus Max (Annual)": {"type": "STRING"},
            "Equity Min (Annual, Amortized)": {"type": "STRING"},
            "Equity Max (Annual, Amortized)": {"type": "STRING"},

            "Visa Sponsorship Evidence / Source": {"type": "STRING"},
            "Relocation Evidence / Source": {"type": "STRING"},
            "Job Source": {"type": "STRING"},
            "Visa Sponsorship Data Type": {"type": "STRING"},
            "Relocation Data Type": {"type": "STRING"},
            "Source Type": {"type": "STRING"},
            "Original / Official Source": {"type": "STRING"},
            "Official Job URL": {"type": "STRING"},
            "Data Source / Evidence": {"type": "STRING"},

            "otherTechnicalSkillsDetected": {
                "type": "ARRAY",
                "items": {"type": "STRING"},
                "description": "Named technologies/languages/tools explicitly mentioned in the posting that are NOT already covered by the fixed skill list above (e.g. TypeScript, Kotlin, DynamoDB, Redis, Kafka). Do not repeat anything already covered. Omit vague terms."
            },

            "confidence": {
                "type": "STRING",
                "enum": ["High", "Medium", "Low"]
            }
        },
        "required": [
            "Job ID",
            "Date Found",
            "Company",
            "Role Title (Raw)",
            "Standardized Role",
            "Level",
            "Country",
            "City",
            "Location Type",
            "Visa Sponsorship",
            "Relocation Support",
            "Currency",
            "Salary Min",
            "Salary Max",
            "Discovery URL",
            "Notes",
            "Exclude From Analysis",
            "skills",
            "Base Comp Min",
            "Base Comp Max",
            "Bonus Min (Annual)",
            "Bonus Max (Annual)",
            "Equity Min (Annual, Amortized)",
            "Equity Max (Annual, Amortized)",
            "Visa Sponsorship Evidence / Source",
            "Relocation Evidence / Source",
            "Job Source",
            "Visa Sponsorship Data Type",
            "Relocation Data Type",
            "Source Type",
            "Original / Official Source",
            "Official Job URL",
            "Data Source / Evidence",
            "otherTechnicalSkillsDetected",
            "confidence"
        ]
    }

    prompt = f"""
You are the extraction/normalization engine for a career-market intelligence workbook.
Never invent facts. If the posting does not state a fact, use Unknown/empty.
For each skill, use Yes only when the posting explicitly requires, lists, or clearly describes it;
No only when there is enough evidence that it is not part of the role; otherwise Unknown.
Salary values must be numeric in the posting's currency and must not be converted here.
Visa and relocation must be based on explicit evidence in the supplied posting.
Standardized Role should be a concise role family such as Software Engineer, Backend Engineer,
SRE, DevOps Engineer, Cloud/Infrastructure Engineer, AI/ML Engineer, Data Engineer,
Systems Engineer, Security Engineer, QA/Automation Engineer, etc.
Separately, list any named technologies/languages/tools explicitly mentioned in the posting that
are NOT already one of the fixed skills above, in otherTechnicalSkillsDetected. Use the specific
proper name (e.g. "TypeScript", "Kotlin", "DynamoDB", "Redis", "Kafka"), not vague phrases.
If the posting mentions none, return an empty array — do not guess or pad this list.

URL: {url}

JOB PAGE TEXT:
{text}
"""

    payload = {
        "contents": [
            {
                "role": "user",
                "parts": [{"text": prompt}]
            }
        ],
        "generationConfig": {
            "temperature": 0,
            "responseMimeType": "application/json",
            "responseSchema": schema
        }
    }
    
    r = requests.post(
        GEMINI_URL,
        headers={
            "x-goog-api-key": key,
            "Content-Type": "application/json",
        },
        json=payload,
        timeout=90,
    )
    
    if not r.ok:
        print("\n========== GEMINI ERROR ==========")
        print("HTTP STATUS:", r.status_code)
        print("RESPONSE:", r.text)
        print("==================================\n")
        r.raise_for_status()
    
    response = r.json()
    
    try:
        raw = response["candidates"][0]["content"]["parts"][0]["text"]
    except (KeyError, IndexError, TypeError) as e:
        raise ValueError(f"Unexpected Gemini response format: {response}") from e
    
    try:
        data = json.loads(raw)
    except json.JSONDecodeError as e:
        raise ValueError(f"Gemini returned invalid JSON: {raw}") from e

    rec = data
    rec["Date Found"] = rec.get("Date Found") or datetime.now().strftime("%Y-%m-%d")
    rec["Discovery URL"] = url
    rec["Job Source"] = rec.get("Job Source") or source_type(url)
    rec["Source Type"] = rec.get("Source Type") or source_type(url)
    rec["_confidence"] = rec.pop("confidence", "Medium")
    rec["_evidence"] = {
        "visa": rec.get("Visa Sponsorship Evidence / Source", ""),
        "relocation": rec.get("Relocation Evidence / Source", ""),
        "overall": rec.get("Data Source / Evidence", "")
    }
    skills = rec.pop("skills", {})
    for s in BOOL_SKILLS:
        rec[s] = skills.get(s, "Unknown")

    other_detected = rec.pop("otherTechnicalSkillsDetected", []) or []
    other_detected = [s for s in other_detected if isinstance(s, str) and s.strip()]
    surfaced = register_detected_skills(other_detected)  # logs to skill_dictionary.json, never touches workbook
    rec["Additional Skills Detected (Unmapped)"] = ", ".join(sorted(set(other_detected)))
    rec["_new_skills_surfaced"] = surfaced  # subset the reviewer should be asked about (Add/Ignore)

    rec["_raw_text"] = text[:8000]
    return rec

def duplicate_check(rec):
    wb=openpyxl.load_workbook(MASTER, read_only=True, data_only=False)
    ws=wb["01 - Job Database"]
    hits=[]
    idx=None
    row_num=1
    for row in ws.iter_rows(min_row=2, values_only=True):
        row_num += 1
        if idx is None:
            # first row read here is the header row (row 2) — build name->index map, then move on
            idx = {h: i for i, h in enumerate(row) if h}
            continue
        def get(name, default_i):
            i = idx.get(name, default_i)
            return row[i] if i < len(row) else None
        company = get("Company", 2)
        jobid = get("Job ID", 0)
        official = get("Official Job URL", 87)
        discovery = get("Discovery URL", 20)
        if rec.get("Official Job URL") and official==rec.get("Official Job URL"): hits.append(row_num)
        elif rec.get("Discovery URL") and discovery==rec.get("Discovery URL"): hits.append(row_num)
        elif rec.get("Job ID") and jobid and str(jobid)==str(rec.get("Job ID")) and company==rec.get("Company"): hits.append(row_num)
    wb.close()
    return sorted(set(hits))

@app.get("/", response_class=HTMLResponse)
def home():
    return HTMLResponse((BASE/"static"/"index.html").read_text(encoding="utf-8"))

@app.post("/api/analyze")
def analyze(req: AnalyzeRequest):
    try:
        print(f"[CHECKPOINT] analyze() called with url={req.url}", flush=True)
        final_url, text = fetch(req.url)
        print(f"[CHECKPOINT] fetch() done, {len(text)} chars, final_url={final_url}", flush=True)
        rec = ai_extract(text, final_url)
        print(f"[CHECKPOINT] ai_extract() done, company={rec.get('Company')}", flush=True)
        rec["_fetched_url"]=final_url
        rec["_duplicate_rows"]=duplicate_check(rec)
        print(f"[CHECKPOINT] duplicate_check() done", flush=True)
        token=str(uuid.uuid4())
        (STAGING/f"{token}.json").write_text(json.dumps(rec,ensure_ascii=False,indent=2),encoding="utf-8")
        print(f"[CHECKPOINT] staged as {token}, returning response", flush=True)
        return {"token":token,"record":rec}
    except Exception as e:
        print("\n========== ANALYZE ERROR ==========")
        print("Details:", e)
        print("===================================\n")
        raise HTTPException(status_code=400, detail=str(e))

@app.post("/api/approve")
def approve(req: ApproveRequest):
    rec=req.record
    # Backup before every write.
    stamp=datetime.now().strftime("%Y%m%d_%H%M%S")
    shutil.copy2(MASTER, BACKUPS/f"Career_Market_Intelligence_V4_{stamp}.xlsx")
    wb=openpyxl.load_workbook(MASTER)
    ws=wb["01 - Job Database"]
    headers=[ws.cell(2,c).value for c in range(1,ws.max_column+1)]
    idx={h:i+1 for i,h in enumerate(headers) if h}
    # Find next empty data row based on Company.
    row=3
    while row <= ws.max_row and ws.cell(row,idx["Company"]).value not in (None,""):
        row += 1
    if row > ws.max_row: row=ws.max_row+1
    for h in headers:
        if not h or h not in rec: continue
        # Never overwrite formula/calculated columns.
        cell=ws.cell(row,idx[h])
        if isinstance(cell.value,str) and cell.value.startswith("="):
            continue
        cell.value=rec[h]
    # Copy formulas from the previous row into calculated cells where appropriate.
    if row > 3:
        for c in range(1,ws.max_column+1):
            prev=ws.cell(row-1,c).value
            cell=ws.cell(row,c)
            if isinstance(prev,str) and prev.startswith("=") and (cell.value in (None,"")):
                # Shift relative references by one row using Translator.
                try:
                    from openpyxl.formula.translate import Translator
                    cell.value=Translator(prev, origin=ws.cell(row-1,c).coordinate).translate_formula(cell.coordinate)
                except Exception:
                    cell.value=prev
    wb.save(MASTER)
    return {"status":"added","row":row,"backup":stamp}

@app.get("/api/master")
def download_master():
    return FileResponse(MASTER, filename=MASTER.name,
                        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

@app.get("/api/skill-dictionary")
def get_skill_dictionary():
    return load_skill_dict()

@app.post("/api/skill-dictionary/add")
def add_skill(req: SkillDictAction):
    """User clicked [Add] on a newly-detected skill. This does NOT touch the
    workbook. It just moves the skill from 'pending' into 'confirmed_for_next_batch'
    so it stops being re-surfaced and is queued for the next promote_skills.py run."""
    name = req.skill.strip()
    if not name:
        raise HTTPException(status_code=400, detail="Empty skill name.")
    d = load_skill_dict()
    d["pending"].pop(name, None)
    if name.lower() not in {s.lower() for s in d["confirmed_for_next_batch"]}:
        d["confirmed_for_next_batch"].append(name)
    save_skill_dict(d)
    return {"status": "confirmed", "skill": name, "confirmed_for_next_batch": d["confirmed_for_next_batch"]}

@app.post("/api/skill-dictionary/ignore")
def ignore_skill(req: SkillDictAction):
    """User clicked [Ignore]. Skill is remembered as ignored so it's never
    surfaced again, but nothing about existing data changes."""
    name = req.skill.strip()
    if not name:
        raise HTTPException(status_code=400, detail="Empty skill name.")
    d = load_skill_dict()
    d["pending"].pop(name, None)
    if name.lower() not in {s.lower() for s in d["ignored"]}:
        d["ignored"].append(name)
    save_skill_dict(d)
    return {"status": "ignored", "skill": name}